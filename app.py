"""NTS Mobile Inspection App - Flask Backend"""
import os, sys, json, uuid, shutil, io, base64
from datetime import datetime
from pathlib import Path
from flask import Flask, request, jsonify, render_template, send_from_directory, session
from functools import wraps

# Add parent to path for report generator
sys.path.insert(0, str(Path(__file__).parent))
from report_generator import build_report, read_workbook

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "nts-inspection-secret-change-me")
APP_PASSWORD = os.environ.get("APP_PASSWORD", "nts2024")
BASE = Path(__file__).parent
INSP_DIR = BASE / "inspections"
PHOTO_DIR = BASE / "photos"
INSP_DIR.mkdir(exist_ok=True)
PHOTO_DIR.mkdir(exist_ok=True)

# ── Database Configuration ──
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL",
    f"sqlite:///{BASE / 'nts_inspections.db'}"
)
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_pre_ping": True,
    "pool_recycle": 300,
}
# Fix Render's postgres:// vs postgresql:// issue
db_url = app.config["SQLALCHEMY_DATABASE_URI"]
if db_url and db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)
    app.config["SQLALCHEMY_DATABASE_URI"] = db_url

from models import db, Inspection, Photo
db.init_app(app)

with app.app_context():
    db.create_all()

# ── Product Templates ─────────────────────────────────────────
PRODUCT_TEMPLATES = {
    "dryer_balls": {
        "name": "Wool Dryer Balls",
        "default_skus": [
            "Penguin and Friends", "Kitty and Friends", "Puppy and Friends",
            "Owl and Friends", "Fox and Friends", "Gorilla and Friends",
            "Hamster and Friends", "Cow and Friends", "Whimsicat and Friends",
            "All Penguins",
        ],
        "packaging_items": [
            "Pouch Fit (6 balls)", "Insert Card in Each Pouch",
            "FNSKU Label Presence", "FBA/Heavy Label on Boxes",
            "Carton Quality & Dimensions",
        ],
        "functional_tests": [
            "Burning Test", "Water Soak Test", "Cut Test",
        ],
        "test_measurements": [
            "Average Pouch Weight (g)", "Ball Weight (g)", "Ball Size (cm)",
        ],
    },
    "singing_bowls": {
        "name": "Singing Bowls",
        "default_skus": [
            "ASB With Box", "ASB Without Box",
            "BSB With Box", "BSB Without Box",
            "BMSB With Box", "BMSB Without Box",
            "BGSB REG With Box", "BGSB REG Without Box",
            "BGSB BRNZ With Box", "BGSB BRNZ Without Box",
            "BAGS With Box", "BAGS Without Box",
            "OBSB With Box", "OBSB Without Box",
            "CHSB With Box", "CHSB Without Box",
            "PTSB With Box", "PTSB Without Box",
        ],
        "packaging_items": [
            "Bubble Wrap & 7-Piece Set Fit", "Unit Packaging Completeness",
            "FNSKU Label Presence", "Heavy Label on Boxes",
            "Carton Quality",
        ],
        "functional_tests": [
            "Sound Test", "Visual Defect Check", "Drop Test",
            "7-Piece Component Check", "Packaging Box Check",
        ],
        "test_measurements": [
            "Package Weight (g)", "Wrap/Box Dimension (cm)",
            "Bowl Diameter (cm)", "FNSKU",
        ],
    },
}

# ── Data Storage ──────────────────────────────────────────────
def load_inspection(insp_id: str) -> dict:
    row = Inspection.query.get(insp_id)
    if row:
        data = dict(row.data)
        data["id"] = row.id
        return data
    return {}

def save_inspection(insp_id: str, data: dict):
    row = Inspection.query.get(insp_id)
    if row:
        row.data = {k: v for k, v in data.items() if k != "id"}
        row.updated_at = datetime.utcnow()
    else:
        row = Inspection(id=insp_id, data={k: v for k, v in data.items() if k != "id"})
        db.session.add(row)
    db.session.commit()

def list_inspections() -> list[dict]:
    return Inspection.list_all()

# ── Auth ───────────────────────────────────────────────────────
def require_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            return jsonify({"error": "Unauthorized"}), 401
        return f(*args, **kwargs)
    return decorated

@app.route("/api/login", methods=["POST"])
def login():
    data = request.json
    if data and data.get("password") == APP_PASSWORD:
        session["logged_in"] = True
        return jsonify({"ok": True})
    return jsonify({"error": "Wrong password"}), 401

@app.route("/api/check-auth")
def check_auth():
    return jsonify({"logged_in": session.get("logged_in", False)})

# ── Routes ────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/templates")
@require_auth
def get_templates():
    return jsonify(PRODUCT_TEMPLATES)

@app.route("/api/inspections")
@require_auth
def get_inspections():
    return jsonify(list_inspections())

@app.route("/api/inspection/new", methods=["POST"])
@require_auth
def new_inspection():
    data = request.json
    insp_id = str(uuid.uuid4())[:8]
    product_type = data.get("product_type", "dryer_balls")
    template = PRODUCT_TEMPLATES.get(product_type, PRODUCT_TEMPLATES["dryer_balls"])

    inspection = {
        "id": insp_id,
        "status": "draft",
        "product_type": product_type,
        "product_name": template["name"],
        "created": datetime.now().isoformat(),
        "updated": datetime.now().isoformat(),
        "client": "",
        "destination": "",
        "supplier": "",
        "factory_address": "Mandev Marg, Byasi-10, Bhaktapur, Nepal",
        "sku_designs": "",
        "inspection_date": "",
        "inspection_duration": "",
        "order_quantity": "",
        "sampled_quantity": "",
        "inspection_location": "On-site at ",
        "report_id": "",
        "inspectors": "Kaustuv Guragain, Amrit Kunwar",
        "aql_standard": "ANSI/ASQ Z1.4-2008",
        "aql_sampling_plan": "Normal, Single",
        "aql_level": "II",
        "aql_critical": "Not Allowed",
        "aql_major": "2.5",
        "aql_minor": "4.5",
        "summary": {
            "A": {"status": "passed", "findings": ""},
            "B": {"status": "passed", "findings": ""},
            "C": {"status": "passed", "findings": ""},
            "D": {"status": "passed", "findings": ""},
            "E": {"status": "passed", "findings": ""},
            "F": {"status": "passed", "findings": ""},
            "G": {"status": "passed", "findings": ""},
            "H": {"status": "passed", "findings": ""},
        },
        "skus": [{"name": s, "ordered": "", "found": "", "missing": "",
                   "sampled": "", "measurements": {}, "defects": "", "ref_category": "",
                   "photos": []} for s in template["default_skus"]],
        "packaging": [{"item": i, "findings": "", "status": "pass"} for i in template["packaging_items"]],
        "functional_tests": [{"test": t, "performed": "yes", "result": "", "status": "pass"} for t in template["functional_tests"]],
        "factory_review": {
            "cooperation": "",
            "workers": "",
            "opinion": "",
        },
        "conclusion": "",
        "media_links": [],
    }
    save_inspection(insp_id, inspection)
    return jsonify({"id": insp_id})

@app.route("/api/inspection/<insp_id>")
@require_auth
def get_inspection(insp_id: str):
    data = load_inspection(insp_id)
    if not data:
        return jsonify({"error": "Not found"}), 404
    return jsonify(data)

@app.route("/api/inspection/<insp_id>/save", methods=["POST"])
@require_auth
def save_inspection_data(insp_id: str):
    data = load_inspection(insp_id)
    if not data:
        return jsonify({"error": "Not found"}), 404
    updates = request.json
    for k, v in updates.items():
        if k in data or k == "status":
            data[k] = v
    data["updated"] = datetime.now().isoformat()
    save_inspection(insp_id, data)
    return jsonify({"ok": True})

@app.route("/api/inspection/<insp_id>/photo", methods=["POST"])
@require_auth
def upload_photo(insp_id: str):
    data = load_inspection(insp_id)
    if not data:
        return jsonify({"error": "Not found"}), 404

    sku_idx = request.form.get("sku_idx", type=int)
    photo_b64 = request.form.get("photo")
    if not photo_b64:
        return jsonify({"error": "No photo data"}), 400

    if "," in photo_b64:
        photo_b64 = photo_b64.split(",")[1]

    # Resize to ~800px before storing
    try:
        from PIL import Image
        import io as io_module
        img_bytes = base64.b64decode(photo_b64)
        img = Image.open(io_module.BytesIO(img_bytes))
        if img.width > 800:
            r = 800 / img.width
            img = img.resize((800, int(img.height * r)), Image.LANCZOS)
        buf = io_module.BytesIO()
        img.save(buf, "JPEG", quality=80)
        photo_b64 = base64.b64encode(buf.getvalue()).decode()
    except:
        pass

    photo_id = f"{insp_id}_{sku_idx}_{datetime.now().strftime('%Y%m%d%H%M%S')}.jpg"
    caption = request.form.get("caption", "")

    # Store in database
    photo = Photo(inspection_id=insp_id, sku_idx=sku_idx,
                  filename=photo_id, caption=caption, data=photo_b64)
    db.session.add(photo)
    db.session.commit()

    # Update inspection data with reference
    if sku_idx is not None and 0 <= sku_idx < len(data.get("skus", [])):
        rel_path = f"photos/{photo_id}"
        photo_entry = {"path": rel_path, "caption": caption}
        data["skus"][sku_idx].setdefault("photos", []).append(photo_entry)
        save_inspection(insp_id, data)

    return jsonify({"ok": True, "photo_id": photo_id})

@app.route("/api/inspection/<insp_id>/generate", methods=["POST"])
@require_auth
def generate_report(insp_id: str):
    data = load_inspection(insp_id)
    if not data:
        return jsonify({"error": "Not found"}), 404

    try:
        # Generate Excel from inspection data
        excel_path = generate_excel(data)
        if not excel_path:
            return jsonify({"error": "Failed to create Excel"}), 500

        # Dump photos from DB to disk for PDF generator
        photo_specs = {}
        db_photos = Photo.query.filter_by(inspection_id=insp_id).all()
        for p in db_photos:
            fpath = PHOTO_DIR / p.filename
            img_bytes = base64.b64decode(p.data)
            fpath.write_bytes(img_bytes)
            photo_specs[str(fpath)] = p.caption or ""

        # Generate PDF (pass photo captions)
        output = build_report(excel_path, PHOTO_DIR, photo_specs)
        
        data["status"] = "completed"
        data["updated"] = datetime.now().isoformat()
        save_inspection(insp_id, data)

        rel_path = Path(output).relative_to(BASE).as_posix()
        return jsonify({"ok": True, "pdf": rel_path})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/inspection/<insp_id>/delete", methods=["POST"])
@require_auth
def delete_inspection(insp_id: str):
    row = Inspection.query.get(insp_id)
    if row:
        db.session.delete(row)
    Photo.query.filter_by(inspection_id=insp_id).delete()
    db.session.commit()
    for f in INSP_DIR.glob(f"{insp_id}_data.xlsx"):
        f.unlink()
    for f in INSP_DIR.glob(f"NTS_Report_{insp_id}.pdf"):
        f.unlink()
    return jsonify({"ok": True})

@app.route("/photos/<path:filename>")
@require_auth
def serve_photo(filename):
    # Try filesystem first, then database
    photo_path = PHOTO_DIR / filename
    if photo_path.exists():
        return send_from_directory(str(PHOTO_DIR), filename)
    # Serve from database
    photo = Photo.query.filter_by(filename=filename).first()
    if photo:
        img_bytes = base64.b64decode(photo.data)
        from flask import Response
        return Response(img_bytes, mimetype="image/jpeg")
    return jsonify({"error": "Not found"}), 404

@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"ok": True})

@app.route("/download/<path:filename>")
@require_auth
def download(filename):
    return send_from_directory(str(BASE), filename, as_attachment=True)

# ── Excel Generator from Inspection Data ──────────────────────
def generate_excel(data: dict) -> Path | None:
    """Create a temporary Excel workbook from inspection data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    
    # Page 2: General Info + Summary
    ws = wb.active
    ws.title = "Page 2"
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 60

    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    r = 6
    ws.cell(r, 1, "I.  GENERAL INFORMATION").font = Font(bold=True, size=11)
    r += 1
    for field, label in [("client", "Client:"), ("destination", "Destination:"),
                          ("supplier", "Supplier:"), ("factory_address", "Factory Address:"),
                          ("product_name", "Product Name:"), ("sku_designs", "SKU Designs:"),
                          ("inspection_date", "Inspection Date:"), ("inspection_duration", "Inspection Duration:"),
                          ("order_quantity", "Order Quantity:"), ("sampled_quantity", "Sampled Quantity:"),
                          ("inspection_location", "Inspection Location:"), ("report_id", "Report ID:"),
                          ("inspectors", "Inspectors:")]:
        ws.cell(r, 1, label).font = bold
        ws.cell(r, 3, data.get(field, ""))
        r += 1

    r += 1
    ws.cell(r, 1, "II.  INSPECTION SUMMARY").font = Font(bold=True, size=11)
    r += 1
    ws.cell(r, 1, "Category"); ws.cell(r, 3, "Passed"); ws.cell(r, 4, "Failed")
    ws.cell(r, 5, "Pending"); ws.cell(r, 6, "Notes / Findings Summary")
    for c in range(1, 7):
        ws.cell(r, c).font = bold
    r += 1

    categories = {
        "A": "A. Quantity", "B": "B. Shape and Size Consistency",
        "C": "C. Workmanship", "D": "D. Product Specification",
        "E": "E. Packing", "F": "F. Marking & Labeling",
        "G": "G. Client Special Requirements", "H": "H. Factory Review",
    }
    for key, label in categories.items():
        s = data.get("summary", {}).get(key, {})
        ws.cell(r, 1, label)
        if s.get("status") == "passed":
            ws.cell(r, 3, "PASSED")
        elif s.get("status") == "failed":
            ws.cell(r, 4, "FAILED")
        ws.cell(r, 6, s.get("findings", "")).alignment = wrap
        r += 2

    # AQL
    ws.cell(r, 1, "Inspection Standard:").font = bold
    ws.cell(r, 4, data.get("aql_standard", ""))
    ws.cell(r, 8, "Critical"); ws.cell(r, 9, "Major"); ws.cell(r, 10, "Minor")
    for c in [8, 9, 10]: ws.cell(r, c).font = bold
    r += 1
    ws.cell(r, 1, "Sampling Plan:").font = bold
    ws.cell(r, 4, data.get("aql_sampling_plan", ""))
    ws.cell(r, 8, data.get("aql_critical", ""))
    ws.cell(r, 9, data.get("aql_major", ""))
    ws.cell(r, 10, data.get("aql_minor", ""))
    r += 1
    ws.cell(r, 1, "Inspection Level:").font = bold
    ws.cell(r, 4, data.get("aql_level", ""))

    # SKU Pages
    for idx, sku in enumerate(data.get("skus", [])):
        ws = wb.create_sheet(f"Page {idx + 3}")
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25

        ws.cell(6, 1, f"III.  SKU WISE INSPECTION: {chr(97 + idx)}").font = Font(bold=True, size=11)

        fields = [
            ("SKU Name", sku.get("name", "")),
            ("Ordered Quantity", sku.get("ordered", "")),
            ("Found in Warehouse", sku.get("found", "")),
            ("Missing Quantity", sku.get("missing", "")),
            ("Sampled Units", sku.get("sampled", "")),
        ]
        for m_key, m_val in sku.get("measurements", {}).items():
            fields.append((m_key, m_val))

        fields.append(("FNSKU", sku.get("fnsku", "")))
        fields.append(("Defects/Findings", sku.get("defects", "")))
        fields.append(("Reference Categories", sku.get("ref_category", "")))

        r = 7
        for label, val in fields:
            ws.cell(r, 1, label).font = bold
            ws.cell(r, 2, val)
            r += 1

    # Last page: Packaging + Tests + Factory Review + Conclusion
    ws_last = wb.create_sheet(f"Page {len(data.get('skus', [])) + 3}")
    ws_last.column_dimensions["A"].width = 40
    ws_last.column_dimensions["C"].width = 60
    ws_last.column_dimensions["D"].width = 15

    r = 6
    ws_last.cell(r, 1, "IV.  PACKAGING, MARKING AND LABELING").font = Font(bold=True, size=11)
    r += 1
    ws_last.cell(r, 1, "Item"); ws_last.cell(r, 3, "Findings"); ws_last.cell(r, 4, "Status/Action")
    for c in [1, 3, 4]: ws_last.cell(r, c).font = bold
    r += 1
    for pkg in data.get("packaging", []):
        ws_last.cell(r, 1, pkg.get("item", ""))
        ws_last.cell(r, 3, pkg.get("findings", "")).alignment = wrap
        ws_last.cell(r, 4, pkg.get("status", "")).alignment = Alignment(horizontal="center")
        r += 1

    r += 1
    ws_last.cell(r, 1, "V.  Functional Tests").font = Font(bold=True, size=11)
    r += 1
    ws_last.cell(r, 1, "Test/Check"); ws_last.cell(r, 2, "Performed")
    ws_last.cell(r, 3, "Result"); ws_last.cell(r, 4, "Notes")
    for c in [1, 2, 3, 4]: ws_last.cell(r, c).font = bold
    r += 1
    for ft in data.get("functional_tests", []):
        ws_last.cell(r, 1, ft.get("test", ""))
        ws_last.cell(r, 2, ft.get("performed", ""))
        ws_last.cell(r, 3, ft.get("result", "")).alignment = wrap
        ws_last.cell(r, 4, ft.get("status", ""))
        r += 1

    r += 2
    ws_last.cell(r, 1, "VI. Factory Review").font = Font(bold=True, size=11)
    r += 1
    fr = data.get("factory_review", {})
    ws_last.cell(r, 1, "Factory cooperation:").font = bold
    ws_last.cell(r, 3, fr.get("cooperation", ""))
    r += 1
    ws_last.cell(r, 1, "Number of workers in factory:").font = bold
    ws_last.cell(r, 3, fr.get("workers", ""))
    r += 1
    ws_last.cell(r, 1, "Inspector's opinion on the Shipment:").font = bold
    ws_last.cell(r, 3, fr.get("opinion", "")).alignment = wrap

    r += 1
    ws_last.cell(r, 1, "VI-A. Media Links").font = Font(bold=True, size=11)
    r += 1
    ws_last.cell(r, 1, "Description"); ws_last.cell(r, 3, "Link")
    for c in [1, 3]: ws_last.cell(r, c).font = bold
    r += 1
    for i, link in enumerate(data.get("media_links", [])):
        if link:
            if isinstance(link, dict):
                label = link.get("label") or f"Media {i+1}"
                url = link.get("url", "")
            else:
                media_labels = ["Sound Tests", "Warehouse & Inspection Clips", "Drop Test Clips", "Product Images"]
                label = media_labels[i] if i < len(media_labels) else f"Media {i+1}"
                url = link
            ws_last.cell(r, 1, label)
            ws_last.cell(r, 3, url)
            r += 1

    r += 2
    ws_last.cell(r, 1, "VII. Conclusion").font = Font(bold=True, size=11)
    r += 1
    ws_last.cell(r, 1, data.get("conclusion", "")).alignment = wrap

    # Save
    excel_path = INSP_DIR / f"{data['id']}_data.xlsx"
    wb.save(str(excel_path))
    return excel_path


# ── Run ───────────────────────────────────────────────────────
if __name__ == "__main__":
    import socket
    host = socket.gethostbyname(socket.gethostname())
    print(f"\n  NTS Inspection App")
    print(f"  Access from your phone at: http://{host}:5000")
    print(f"  Make sure your phone is on the same WiFi network\n")
    app.run(host="0.0.0.0", port=5000, debug=True)
