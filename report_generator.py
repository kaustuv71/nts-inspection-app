"""NTS Report Generator - Read Excel data and produce branded PDF reports."""
import os, sys, math
from pathlib import Path
from datetime import datetime

def read_workbook(excel_path: str) -> dict:
    """Read the intermediate Excel workbook and return parsed data."""
    from openpyxl import load_workbook
    wb = load_workbook(excel_path, data_only=True)

    data = {
        "general": {},
        "summary": {},
        "aql": {},
        "skus": [],
        "packaging": [],
        "functional_tests": [],
        "factory_review": {},
        "conclusion": "",
    }

    # ── Parse Page 2 (General + Summary + AQL) ──
    ws = wb["Page 2"]
    general_fields = [
        ("Client:", "client"), ("Destination:", "destination"),
        ("Supplier:", "supplier"), ("Factory Address:", "factory_address"),
        ("Product Name:", "product_name"), ("SKU Designs:", "sku_designs"),
        ("Inspection Date:", "inspection_date"), ("Inspection Duration:", "inspection_duration"),
        ("Order Quantity:", "order_quantity"), ("Sampled Quantity:", "sampled_quantity"),
        ("Inspection Location:", "inspection_location"), ("Report ID:", "report_id"),
        ("Inspectors:", "inspectors"),
    ]
    for row in ws.iter_rows(min_row=6, max_row=30, values_only=False):
        label_cell = row[0]
        val_cell = row[2] if len(row) > 2 else None
        if label_cell and label_cell.value:
            label = str(label_cell.value).strip()
            for prefix, key in general_fields:
                if label.startswith(prefix):
                    data["general"][key] = str(val_cell.value or "").strip()
                    break

    # Summary categories (rows after "II. INSPECTION SUMMARY")
    cat_keys = ["A", "B", "C", "D", "E", "F", "G", "H"]
    cat_labels = [
        "A. Quantity", "B. Shape and Size Consistency",
        "C. Workmanship", "D. Product Specification",
        "E. Packing", "F. Marking & Labeling",
        "G. Client Special Requirements", "H. Factory Review",
    ]
    for row in ws.iter_rows(min_row=34, max_row=60, values_only=False):
        label_val = str(row[0].value or "").strip() if row[0] else ""
        for ck, cl in zip(cat_keys, cat_labels):
            if label_val.startswith(cl.replace(".", "")) or label_val == cl:
                status = "passed"
                if row[3] and str(row[3].value).strip().upper() == "FAILED":
                    status = "failed"
                elif row[4] and str(row[4].value).strip().upper() == "FAILED":
                    status = "failed"
                findings = str(row[5].value or "") if len(row) > 5 and row[5] else ""
                data["summary"][ck] = {"status": status, "findings": findings}
                break

    # AQL
    for row in ws.iter_rows(min_row=40, max_row=55, values_only=False):
        label_val = str(row[0].value or "").strip() if row[0] else ""
        if "Inspection Standard" in label_val or "Standard:" in label_val:
            data["aql"]["standard"] = str(row[3].value or "").strip() if len(row) > 3 else ""
            data["aql"]["critical"] = str(row[7].value or "").strip() if len(row) > 7 else ""
            data["aql"]["major"] = str(row[8].value or "").strip() if len(row) > 8 else ""
            data["aql"]["minor"] = str(row[9].value or "").strip() if len(row) > 9 else ""
        elif "Sampling" in label_val:
            data["aql"]["sampling_plan"] = str(row[3].value or "").strip() if len(row) > 3 else ""
        elif "Level" in label_val:
            data["aql"]["level"] = str(row[3].value or "").strip() if len(row) > 3 else ""

    # ── Parse SKU Pages ──
    for ws_name in wb.sheetnames:
        if ws_name.startswith("Page ") and ws_name != "Page 2":
            ws = wb[ws_name]
            sku = {"name": f"SKU {len(data['skus']) + 1}", "fields": {}}
            for row in ws.iter_rows(min_row=6, max_row=25, values_only=False):
                label_val = str(row[0].value or "").strip() if row[0] else ""
                val = str(row[1].value or "").strip() if len(row) > 1 and row[1] else ""
                if "SKU Name" in label_val or "name" in label_val.lower():
                    sku["name"] = val if val else sku["name"]
                elif label_val and val:
                    sku["fields"][label_val] = val
            data["skus"].append(sku)

    # ── Parse Last page ──
    last_page_name = f"Page {len(data['skus']) + 3}"
    if last_page_name in wb.sheetnames:
        ws = wb[last_page_name]
        section = None
        for row in ws.iter_rows(min_row=6, max_row=50, values_only=False):
            label_val = str(row[0].value or "").strip() if row[0] else ""
            val3 = str(row[2].value or "").strip() if len(row) > 2 and row[2] else ""
            val4 = str(row[3].value or "").strip() if len(row) > 3 and row[3] else ""

            if "PACKAGING" in label_val.upper():
                section = "packaging"
                continue
            elif "TESTS" in label_val.upper() or "FUNCTIONAL" in label_val.upper():
                section = "functional_tests"
                continue
            elif "MEDIA" in label_val.upper():
                section = "media"
                continue
            elif "FACTORY REVIEW" in label_val.upper():
                section = "factory"
                continue
            elif "CONCLUSION" in label_val.upper():
                section = "conclusion"
                continue

            if section == "packaging" and label_val and label_val != "Item":
                data["packaging"].append({
                    "item": label_val,
                    "findings": val3,
                    "status": val4.lower() if val4 else "pass",
                })
            elif section == "functional_tests" and label_val and label_val != "Test/Check":
                data["functional_tests"].append({
                    "test": label_val,
                    "performed": val3,
                    "result": val4,
                    "status": "pass",
                })
            elif section == "media" and val3:
                data.setdefault("media_links", []).append(val3)
            elif section == "factory" and label_val:
                if "cooperation" in label_val.lower():
                    data["factory_review"]["cooperation"] = val3
                elif "workers" in label_val.lower():
                    data["factory_review"]["workers"] = val3
                elif "opinion" in label_val.lower():
                    data["factory_review"]["opinion"] = val3
            elif section == "conclusion" and label_val:
                data["conclusion"] = label_val

    return data


def build_report(excel_path: str, photo_dir: str, photo_specs: dict | None = None) -> str:
    """Build a professional PDF report from the Excel workbook and photos.

    photo_specs: dict mapping photo path -> caption string (for defect descriptions in red).
    Returns the path to the generated PDF.
    """
    from fpdf import FPDF

    data = read_workbook(excel_path)
    g = data.get("general", {})
    insp_id = Path(excel_path).stem.replace("_data", "")
    output_dir = Path(excel_path).parent
    pdf_path = output_dir / f"NTS_Report_{insp_id}.pdf"

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.alias_nb_pages()

    # ── Color Constants ──
    NAVY = (15, 30, 55)
    GOLD = (198, 156, 75)
    WHITE = (255, 255, 255)
    LIGHT_GRAY = (245, 243, 238)
    DARK_TEXT = (40, 40, 40)
    GREEN = (21, 87, 36)
    RED = (114, 28, 36)
    PAGE_W = 210
    MARGIN = 15
    CONTENT_W = PAGE_W - 2 * MARGIN

    logo_path = Path(__file__).parent / "nts_logo.png"

    def footer_block():
        pdf.set_fill_color(*NAVY)
        pdf.rect(0, pdf.h - 12, 210, 12, "F")
        pdf.set_y(pdf.h - 10)
        pdf.set_text_color(*GOLD)
        pdf.set_font("Helvetica", "", 7)
        pdf.cell(0, 5, f"NTS Report | Page {pdf.page_no()}/{{nb}}", align="C")

    def header_block():
        pdf.set_fill_color(*NAVY)
        pdf.rect(0, 0, 210, 30, "F")
        if logo_path.exists():
            pdf.image(str(logo_path), x=8, y=3, w=24)
        rid = g.get("report_id", "")
        pdf.set_text_color(*GOLD)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_xy(160, 5)
        pdf.cell(40, 5, rid, ln=True, align="R")
        pdf.set_font("Helvetica", "B", 14)
        pdf.set_xy(38, 5)
        pdf.cell(115, 5, "NEPAL TRADE SOLUTIONS", ln=True, align="L")
        pdf.set_font("Helvetica", "", 7)
        pdf.set_text_color(*WHITE)
        pdf.set_xy(38, 12)
        pdf.cell(115, 4, "Quality Inspection Report | Mandev Marg, Byasi-10, Bhaktapur, Nepal", ln=True, align="L")
        pdf.set_font("Helvetica", "I", 6)
        pdf.set_xy(38, 18)
        pdf.cell(115, 4, "www.nepalts.com", ln=True, align="L")

    def section_title(num, title, y_adj=20):
        pdf.ln(y_adj)
        pdf.set_text_color(*NAVY)
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, f"{num}. {title}", ln=True)
        pdf.set_draw_color(*GOLD)
        pdf.line(MARGIN, pdf.get_y(), PAGE_W - MARGIN, pdf.get_y())
        pdf.ln(5)

    # ────────────────────────────────────────────────────
    # COVER PAGE
    # ────────────────────────────────────────────────────
    pdf.add_page()
    header_block()
    footer_block()

    if logo_path.exists():
        pdf.ln(18)
        pdf.image(str(logo_path), x=65, y=pdf.get_y(), w=80)
        pdf.ln(36)
    else:
        pdf.ln(40)

    pdf.set_text_color(*NAVY)
    pdf.set_font("Helvetica", "B", 28)
    pdf.cell(0, 15, "INSPECTION REPORT", ln=True, align="C")

    report_id = g.get("report_id", "")
    if report_id:
        pdf.ln(3)
        pdf.set_text_color(*GOLD)
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, report_id, ln=True, align="C")

    # Cover fields in a structured card
    pdf.ln(6)
    cover_fields = [
        ("Client", g.get("client", "-")),
        ("Product", g.get("product_name", "-")),
        ("Supplier", g.get("supplier", "-")),
        ("Inspection Date", g.get("inspection_date", "-")),
        ("Inspectors", g.get("inspectors", "-")),
    ]
    card_x = 45
    card_w = 120
    field_h = 8
    card_h = len(cover_fields) * field_h + 6
    card_y = pdf.get_y()

    pdf.set_fill_color(*LIGHT_GRAY)
    pdf.rect(card_x, card_y, card_w, card_h, "F")
    pdf.set_fill_color(*GOLD)
    pdf.rect(card_x, card_y, 2, card_h, "F")

    pdf.set_xy(card_x + 8, card_y + 3)
    for label, val in cover_fields:
        pdf.set_x(card_x + 8)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(*NAVY)
        pdf.cell(28, field_h, label + ":")
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*DARK_TEXT)
        pdf.cell(0, field_h, val, ln=True)

    pdf.set_y(card_y + card_h + 12)

    pdf.set_text_color(*GOLD)
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 8, "NEPAL TRADE SOLUTIONS", ln=True, align="C")
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(*DARK_TEXT)
    pdf.cell(0, 5, "Mandev Marg, Byasi-10, Bhaktapur, Nepal", ln=True, align="C")
    pdf.ln(4)
    pdf.set_fill_color(*GOLD)
    pdf.rect(60, pdf.get_y(), 90, 0.5, "F")
    pdf.ln(5)
    pdf.set_text_color(*NAVY)
    pdf.set_font("Helvetica", "I", 8)
    pdf.cell(0, 5, "This report is prepared solely for the use of the client named above", ln=True, align="C")
    pdf.cell(0, 5, "and contains confidential information.", ln=True, align="C")

    # ────────────────────────────────────────────────────
    # TABLE OF CONTENTS
    # ────────────────────────────────────────────────────
    pdf.add_page()
    header_block()
    footer_block()

    section_title("", "TABLE OF CONTENTS", 20)
    toc_items = [
        ("I", "General Information"),
        ("II", "Inspection Summary & Results"),
        ("III", "SKU-Wise Inspection Details"),
        ("IV", "Packaging, Marking & Labeling"),
        ("V", "Functional Tests"),
        ("VI", "Media & Attachments"),
        ("VII", "Factory Review"),
        ("VIII", "Conclusion & Signatures"),
    ]
    pdf.set_text_color(*DARK_TEXT)
    for num, title in toc_items:
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(10, 8, num, ln=False)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(140, 8, title, ln=False)
        pdf.set_text_color(*GOLD)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(0, 8, "...", ln=True, align="R")
        pdf.set_text_color(*DARK_TEXT)
        pdf.ln(0)

    # ────────────────────────────────────────────────────
    # PAGE: GENERAL INFORMATION
    # ────────────────────────────────────────────────────
    pdf.add_page()
    header_block()
    footer_block()

    section_title("I", "GENERAL INFORMATION")
    rows = [
        ("Client:", g.get("client", "")),
        ("Destination:", g.get("destination", "")),
        ("Supplier:", g.get("supplier", "")),
        ("Factory Address:", g.get("factory_address", "")),
        ("Product Name:", g.get("product_name", "")),
        ("SKU Designs:", g.get("sku_designs", "")),
        ("Inspection Date:", g.get("inspection_date", "")),
        ("Duration:", g.get("inspection_duration", "")),
        ("Order Quantity:", g.get("order_quantity", "")),
        ("Sampled Quantity:", g.get("sampled_quantity", "")),
        ("Location:", g.get("inspection_location", "")),
        ("Report ID:", g.get("report_id", "")),
        ("Inspectors:", g.get("inspectors", "")),
    ]
    pdf.set_text_color(*DARK_TEXT)
    for i, (label, val) in enumerate(rows):
        y = pdf.get_y()
        if i % 2 == 1:
            pdf.set_fill_color(*LIGHT_GRAY)
            pdf.rect(MARGIN, y, CONTENT_W, 7, "F")
        pdf.set_xy(MARGIN + 2, y)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(42, 7, label)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(0, 7, val, ln=True)
        if label == "Factory Address:":
            pdf.ln(1)

    # ────────────────────────────────────────────────────
    # PAGE: SUMMARY + EXECUTIVE SUMMARY + CHART
    # ────────────────────────────────────────────────────
    pdf.add_page()
    header_block()
    footer_block()

    # Executive Summary Box
    summary_data = data.get("summary", {})
    cat_count = len(summary_data)
    passed = sum(1 for s in summary_data.values() if s.get("status") == "passed")
    failed = sum(1 for s in summary_data.values() if s.get("status") == "failed")
    pending = cat_count - passed - failed
    pass_rate = round((passed / cat_count * 100)) if cat_count > 0 else 0

    pdf.ln(12)
    box_y = pdf.get_y()

    pdf.set_fill_color(*LIGHT_GRAY)
    pdf.rect(MARGIN, box_y, CONTENT_W, 30, "F")
    pdf.set_fill_color(*GOLD)
    pdf.rect(MARGIN, box_y, CONTENT_W, 2, "F")

    pdf.set_xy(MARGIN + 5, box_y + 5)
    pdf.set_text_color(*NAVY)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 6, "EXECUTIVE SUMMARY", ln=True)

    pdf.set_xy(MARGIN + 5, box_y + 13)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(*DARK_TEXT)
    pdf.cell(40, 5, f"Categories Inspected:  {cat_count}")
    pdf.set_text_color(*GREEN)
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(30, 5, f"Passed:  {passed}")
    pdf.set_text_color(*RED)
    pdf.cell(30, 5, f"Failed:  {failed}")
    pdf.set_text_color(*NAVY)
    pdf.cell(30, 5, f"Pending:  {pending}")

    pdf.set_xy(MARGIN + 5, box_y + 20)
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(*NAVY)
    pass_label = "PASS" if pass_rate >= 80 else "CONDITIONAL" if pass_rate >= 50 else "FAIL"
    pdf.cell(0, 5, f"Overall Result:  {pass_label}  ({pass_rate}% pass rate)", ln=True)

    pdf.set_y(box_y + 35)

    # Section II title
    section_title("II", "INSPECTION SUMMARY", 1)
    y_before_chart = pdf.get_y()

    # Donut chart on the right
    chart_cx = PAGE_W - MARGIN - 30
    chart_cy = y_before_chart + 15
    outer_r = 14
    inner_r = 6

    if cat_count > 0 and failed > 0:
        failed_angle = (failed / cat_count) * 360
        passed_angle = (passed / cat_count) * 360

        pdf.set_fill_color(*RED)
        pdf.set_draw_color(*RED)
        pts = [(chart_cx, chart_cy)]
        for deg in range(0, int(failed_angle) + 1, 2):
            pts.append((
                chart_cx + outer_r * math.cos(math.radians(deg - 90)),
                chart_cy + outer_r * math.sin(math.radians(deg - 90))
            ))
        if len(pts) > 2:
            pdf.polygon(pts, style="DF")

        pdf.set_fill_color(*GREEN)
        pdf.set_draw_color(*GREEN)
        pts = [(chart_cx, chart_cy)]
        for deg in range(0, int(passed_angle) + 1, 2):
            pts.append((
                chart_cx + outer_r * math.cos(math.radians(deg - 90 + failed_angle)),
                chart_cy + outer_r * math.sin(math.radians(deg - 90 + failed_angle))
            ))
        if len(pts) > 2:
            pdf.polygon(pts, style="DF")
    else:
        pdf.set_fill_color(*GREEN)
        pdf.set_draw_color(*GREEN)
        pts = [(chart_cx, chart_cy)]
        for deg in range(0, 361, 2):
            pts.append((
                chart_cx + outer_r * math.cos(math.radians(deg - 90)),
                chart_cy + outer_r * math.sin(math.radians(deg - 90))
            ))
        pdf.polygon(pts, style="DF")

    # Donut hole
    pdf.set_fill_color(*WHITE)
    pdf.set_draw_color(*WHITE)
    hole_pts = []
    for deg in range(0, 361, 2):
        hole_pts.append((
            chart_cx + inner_r * math.cos(math.radians(deg - 90)),
            chart_cy + inner_r * math.sin(math.radians(deg - 90))
        ))
    pdf.polygon(hole_pts, style="DF")

    pdf.set_text_color(*NAVY)
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_xy(chart_cx - 5, chart_cy - 3)
    pdf.cell(10, 5, f"{pass_rate}%", align="C")

    leg_x = chart_cx - 12
    leg_y = chart_cy + outer_r + 6
    pdf.set_fill_color(*GREEN)
    pdf.rect(leg_x, leg_y, 4, 4, "F")
    pdf.set_font("Helvetica", "", 6)
    pdf.set_text_color(*DARK_TEXT)
    pdf.set_xy(leg_x + 6, leg_y - 1)
    pdf.cell(10, 4, f"Pass ({passed})")
    pdf.set_fill_color(*RED)
    pdf.rect(leg_x, leg_y + 5, 4, 4, "F")
    pdf.set_xy(leg_x + 6, leg_y + 4)
    pdf.cell(10, 4, f"Fail ({failed})")

    # Summary table
    pdf.set_xy(MARGIN, y_before_chart)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(*NAVY)
    pdf.set_text_color(*WHITE)
    pdf.cell(50, 7, "Category", 1, 0, "C", True)
    pdf.cell(12, 7, "Status", 1, 0, "C", True)
    pdf.cell(0, 7, "Notes / Findings Summary", 1, 1, "C", True)

    cat_titles = {
        "A": "A. Quantity",
        "B": "B. Shape & Size Consistency",
        "C": "C. Workmanship",
        "D": "D. Product Specification",
        "E": "E. Packing",
        "F": "F. Marking & Labeling",
        "G": "G. Client Requirements",
        "H": "H. Factory Review",
    }
    pdf.set_text_color(*DARK_TEXT)
    for row_idx, (ck, cl) in enumerate(cat_titles.items()):
        s = data.get("summary", {}).get(ck, {})
        status = s.get("status", "pending").upper()
        findings = s.get("findings", "")

        lines_needed = max(1, len(str(findings)) // 55 + 1)
        row_h = max(6, lines_needed * 4)

        y_before = pdf.get_y()
        if y_before + row_h > 268:
            pdf.add_page()
            header_block()
            footer_block()
            pdf.set_text_color(*NAVY)
            pdf.set_font("Helvetica", "B", 14)
            pdf.ln(16)
            pdf.cell(0, 10, "II. INSPECTION SUMMARY (cont.)", ln=True)
            pdf.set_draw_color(*GOLD)
            pdf.line(MARGIN, pdf.get_y(), PAGE_W - MARGIN, pdf.get_y())
            pdf.ln(4)
            pdf.set_fill_color(*NAVY)
            pdf.set_text_color(*WHITE)
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(50, 7, "Category", 1, 0, "C", True)
            pdf.cell(12, 7, "Status", 1, 0, "C", True)
            pdf.cell(0, 7, "Notes / Findings Summary", 1, 1, "C", True)
            pdf.set_text_color(*DARK_TEXT)
            y_before = pdf.get_y()

        if row_idx % 2 == 1:
            pdf.set_fill_color(*LIGHT_GRAY)
            pdf.rect(MARGIN, y_before, CONTENT_W, row_h, "F")

        pdf.set_xy(MARGIN, y_before)
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(50, row_h, cl, 1, 0, "L")
        pdf.set_font("Helvetica", "B", 8)
        status_color = GREEN if status == "PASSED" else RED
        pdf.set_text_color(*status_color)
        pdf.cell(12, row_h, "PASSED" if status == "PASSED" else "FAILED", 1, 0, "C")
        pdf.set_text_color(*DARK_TEXT)
        pdf.set_font("Helvetica", "", 7)
        pdf.cell(0, row_h, findings, 1, 1, "L")

    # AQL Section
    aql = data.get("aql", {})
    pdf.ln(8)
    pdf.set_text_color(*NAVY)
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "AQL Details (Based on the inspection of the finished products)", ln=True)
    pdf.set_draw_color(*GOLD)
    pdf.line(MARGIN, pdf.get_y(), PAGE_W - MARGIN, pdf.get_y())
    pdf.ln(4)
    pdf.set_text_color(*DARK_TEXT)
    aql_rows = [
        ("Inspection Standard:", aql.get("standard", "ANSI/ASQ Z1.4-2008"), ""),
        ("Sampling Plan:", aql.get("sampling_plan", "Normal, Single"), ""),
        ("Inspection Level:", aql.get("level", "II"), ""),
        ("", "", f"Critical:  {aql.get('critical', 'Not Allowed')}    Major:  {aql.get('major', '2.5')}    Minor:  {aql.get('minor', '4.5')}"),
    ]
    for i, (label, val, extra) in enumerate(aql_rows):
        if i % 2 == 1:
            pdf.set_fill_color(*LIGHT_GRAY)
            pdf.rect(MARGIN, pdf.get_y(), CONTENT_W, 7, "F")
        if label:
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(35, 7, label)
            pdf.set_font("Helvetica", "", 9)
            pdf.cell(0, 7, val, ln=True)
        else:
            pdf.cell(0, 7, extra, ln=True)

    order_qty = g.get("order_quantity", "")
    sample_qty = g.get("sampled_quantity", "")
    for i, (label, val) in enumerate([("Order Quantity:", order_qty), ("Available Quantity:", order_qty),
                                       ("Sample Size:", sample_qty)]):
        if val:
            if (len(aql_rows) + i) % 2 == 1:
                pdf.set_fill_color(*LIGHT_GRAY)
                pdf.rect(MARGIN, pdf.get_y(), CONTENT_W, 7, "F")
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(35, 7, label)
            pdf.set_font("Helvetica", "", 9)
            pdf.cell(0, 7, val, ln=True)

    # ────────────────────────────────────────────────────
    # SKU PAGES
    # ────────────────────────────────────────────────────
    for idx, sku in enumerate(data.get("skus", [])):
        pdf.add_page()
        header_block()
        footer_block()

        sku_name = sku.get("name", f"SKU {idx + 1}")
        pdf.set_text_color(*NAVY)
        pdf.set_font("Helvetica", "B", 14)
        pdf.ln(18)
        pdf.cell(0, 10, "III. SKU INSPECTION", ln=True)
        pdf.set_draw_color(*GOLD)
        pdf.line(MARGIN, pdf.get_y(), PAGE_W - MARGIN, pdf.get_y())
        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(*GOLD)
        pdf.cell(0, 8, f"SKU {idx + 1}:  {sku_name}", ln=True)
        pdf.set_text_color(*DARK_TEXT)
        pdf.ln(2)

        fields = sku.get("fields", {})
        field_order = [
            "Ordered Quantity", "Found in Warehouse", "Missing Quantity",
            "Sampled Units", "FNSKU", "Defects/Findings",
            "Reference Categories",
        ]
        for fi, key in enumerate(field_order):
            val = fields.get(key, "")
            if val:
                if fi % 2 == 1:
                    pdf.set_fill_color(*LIGHT_GRAY)
                    pdf.rect(MARGIN, pdf.get_y(), CONTENT_W, 7, "F")
                pdf.set_font("Helvetica", "B", 9)
                pdf.cell(50, 7, key + ":")
                pdf.set_font("Helvetica", "", 9)
                pdf.cell(0, 7, val, ln=True)

        # Measurements
        meas_keys = [k for k in fields if k not in field_order and k != "SKU Name"]
        if meas_keys:
            pdf.ln(3)
            pdf.set_fill_color(*NAVY)
            pdf.set_text_color(*WHITE)
            pdf.set_font("Helvetica", "B", 9)
            for mi, mk in enumerate(meas_keys):
                pdf.cell(55, 6, "Measurement", 1, 0, "C", True)
                pdf.cell(0, 6, "Value", 1, 1, "C", True)
                pdf.set_text_color(*DARK_TEXT)
                if mi % 2 == 1:
                    pdf.set_fill_color(*LIGHT_GRAY)
                pdf.set_font("Helvetica", "", 9)
                pdf.cell(55, 6, mk, 1)
                pdf.cell(0, 6, fields.get(mk, ""), 1, 1)
                pdf.ln(1)
                pdf.set_text_color(*WHITE)
                pdf.set_font("Helvetica", "B", 9)
            pdf.set_text_color(*DARK_TEXT)

        # Photos
        photos = []
        if photo_dir:
            photo_path = Path(photo_dir)
            if photo_path.exists():
                photos = sorted(photo_path.glob(f"{insp_id}_{idx}_*.jpg"))
                photos += sorted(photo_path.glob(f"{insp_id}_{idx}_*.jpeg"))

        pdf.ln(2)
        if photos:
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(0, 6, "Photos:", ln=True)
            x_start = pdf.get_x()
            y_start = pdf.get_y()
            img_w = 55
            img_h = 40
            gap = 3
            max_per_row = 3
            for pi, photo_file in enumerate(photos):
                col = pi % max_per_row
                row_n = pi // max_per_row
                x = x_start + col * (img_w + gap)
                y = y_start + row_n * (img_h + gap)
                cap_h = 6
                if y + img_h + cap_h > 270:
                    break
                try:
                    pdf.image(str(photo_file), x=x, y=y, w=img_w, h=img_h)
                    fname = Path(photo_file).name
                    if photo_specs:
                        for pkey, cap in photo_specs.items():
                            if fname in pkey or pkey.endswith(fname):
                                if cap:
                                    pdf.set_font("Helvetica", "I", 6)
                                    pdf.set_text_color(200, 30, 30)
                                    pdf.set_xy(x, y + img_h + 1)
                                    pdf.multi_cell(img_w, 3, cap[:80])
                                    pdf.set_text_color(*DARK_TEXT)
                                break
                except:
                    pass
        else:
            pdf.set_font("Helvetica", "I", 8)
            pdf.set_text_color(*GOLD)
            pdf.cell(0, 6, "No photos uploaded for this SKU.", ln=True)
            pdf.set_text_color(*DARK_TEXT)

    # ────────────────────────────────────────────────────
    # PACKAGING + TESTS + FACTORY + CONCLUSION
    # ────────────────────────────────────────────────────
    pdf.add_page()
    header_block()
    footer_block()

    # IV. Packaging
    section_title("IV", "PACKAGING, MARKING & LABELING", 18)

    if data.get("packaging"):
        pdf.set_fill_color(*NAVY)
        pdf.set_text_color(*WHITE)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(55, 7, "Item", 1, 0, "C", True)
        pdf.cell(80, 7, "Findings", 1, 0, "C", True)
        pdf.cell(0, 7, "Status", 1, 1, "C", True)
        pdf.set_text_color(*DARK_TEXT)

        for pi, pkg in enumerate(data["packaging"]):
            pdf.set_font("Helvetica", "", 8)
            findings = pkg.get("findings", "")
            lines = max(1, len(findings) // 45 + 1)
            rh = max(6, lines * 4)
            y_before = pdf.get_y()
            if pi % 2 == 1:
                pdf.set_fill_color(*LIGHT_GRAY)
                pdf.rect(MARGIN, y_before, CONTENT_W, rh, "F")
            pdf.set_xy(MARGIN, y_before)
            pdf.cell(55, rh, pkg.get("item", ""), 1, 0, "L")
            pdf.cell(80, rh, findings, 1, 0, "L")
            status = pkg.get("status", "").upper()
            sc = GREEN if status in ("PASS", "PASSED") else RED
            pdf.set_text_color(*sc)
            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(0, rh, status, 1, 1, "C")
            pdf.set_text_color(*DARK_TEXT)
    else:
        pdf.set_font("Helvetica", "I", 8)
        pdf.cell(0, 6, "No packaging data recorded.", ln=True)

    # V. Functional Tests
    pdf.ln(8)
    section_title("V", "FUNCTIONAL TESTS", 0)

    if data.get("functional_tests"):
        pdf.set_fill_color(*NAVY)
        pdf.set_text_color(*WHITE)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(40, 7, "Test", 1, 0, "C", True)
        pdf.cell(15, 7, "Done", 1, 0, "C", True)
        pdf.cell(90, 7, "Result", 1, 0, "C", True)
        pdf.cell(0, 7, "Status", 1, 1, "C", True)
        pdf.set_text_color(*DARK_TEXT)

        for fi, ft in enumerate(data["functional_tests"]):
            pdf.set_font("Helvetica", "", 8)
            result = ft.get("result", "")
            lines = max(1, len(result) // 50 + 1)
            rh = max(6, lines * 4)
            y_before = pdf.get_y()
            if fi % 2 == 1:
                pdf.set_fill_color(*LIGHT_GRAY)
                pdf.rect(MARGIN, y_before, CONTENT_W, rh, "F")
            pdf.set_xy(MARGIN, y_before)
            pdf.cell(40, rh, ft.get("test", ""), 1, 0, "L")
            pdf.cell(15, rh, ft.get("performed", ""), 1, 0, "C")
            pdf.cell(90, rh, result, 1, 0, "L")
            status = ft.get("status", "").upper()
            sc = GREEN if status in ("PASS", "PASSED") else RED
            pdf.set_text_color(*sc)
            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(0, rh, status, 1, 1, "C")
            pdf.set_text_color(*DARK_TEXT)
    else:
        pdf.set_font("Helvetica", "I", 8)
        pdf.cell(0, 6, "No functional test data recorded.", ln=True)

    # VI. Media Links
    media_links = data.get("media_links", [])
    if media_links and any(media_links):
        pdf.ln(6)
        section_title("VI", "MEDIA & ATTACHMENTS", 0)
        pdf.set_fill_color(*NAVY)
        pdf.set_text_color(*WHITE)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(55, 7, "Description", 1, 0, "C", True)
        pdf.cell(0, 7, "Link", 1, 1, "C", True)
        pdf.set_text_color(*DARK_TEXT)

        for li, link in enumerate(media_links):
            if link:
                if isinstance(link, dict):
                    label = link.get("label") or f"Attachment {li+1}"
                    url = link.get("url", "")
                else:
                    media_labels = ["Sound Tests", "Warehouse & Inspection Clips", "Drop Test Clips", "Product Images"]
                    label = media_labels[li] if li < len(media_labels) else f"Attachment {li+1}"
                    url = link
                y_before = pdf.get_y()
                if li % 2 == 1:
                    pdf.set_fill_color(*LIGHT_GRAY)
                    pdf.rect(MARGIN, y_before, CONTENT_W, 6, "F")
                pdf.set_xy(MARGIN, y_before)
                pdf.set_font("Helvetica", "", 8)
                pdf.cell(55, 6, label, 1, 0, "L")
                pdf.cell(0, 6, url, 1, 1, "L")
        pdf.ln(3)

    # VII. Factory Review
    pdf.ln(4 if (media_links and any(media_links)) else 8)
    section_title("VII", "FACTORY REVIEW", 0)

    fr = data.get("factory_review", {})
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(*DARK_TEXT)
    for fi, (label, val) in enumerate([
        ("Factory Cooperation:", fr.get("cooperation", "-")),
        ("Number of Workers:", fr.get("workers", "-")),
    ]):
        if fi % 2 == 1:
            pdf.set_fill_color(*LIGHT_GRAY)
            pdf.rect(MARGIN, pdf.get_y(), CONTENT_W, 7, "F")
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(40, 7, label)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(0, 7, val, ln=True)

    opinion = fr.get("opinion", "")
    pdf.cell(0, 6, "Inspector's Opinion on Shipment:", ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.multi_cell(0, 5, opinion if opinion else "(No opinion recorded)")

    # VIII. Conclusion
    pdf.ln(8)
    section_title("VIII", "CONCLUSION", 0)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(*DARK_TEXT)
    conclusion = data.get("conclusion", "")
    pdf.multi_cell(0, 5, conclusion if conclusion else "(No conclusion entered)")

    # ────────────────────────────────────────────────────
    # SIGNATURES & STAMP
    # ────────────────────────────────────────────────────
    pdf.ln(15)

    def draw_stamp(x, y, size=36):
        r = size / 2
        cx = x + r
        cy = y + r
        pdf.set_line_width(1.5)
        pdf.set_draw_color(*NAVY)
        pdf.circle(x, y, r)
        pdf.set_draw_color(*GOLD)
        pdf.circle(x + 2, y + 2, r - 4)
        pdf.set_line_width(0.3)
        pdf.set_text_color(*NAVY)
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_xy(cx - 14, cy - 10)
        pdf.cell(28, 3, "NEPAL TRADE", ln=True, align="C")
        pdf.set_xy(cx - 14, cy - 7)
        pdf.cell(28, 3, "SOLUTIONS", ln=True, align="C")
        pdf.set_draw_color(*NAVY)
        pdf.line(cx - 10, cy - 3, cx + 10, cy - 3)
        pdf.set_font("Helvetica", "B", 6)
        pdf.set_xy(cx - 14, cy - 1)
        pdf.cell(28, 3, "INSPECTION", ln=True, align="C")
        pdf.line(cx - 10, cy + 3, cx + 10, cy + 3)
        pdf.set_font("Helvetica", "", 6)
        pdf.set_xy(cx - 14, cy + 5)
        pdf.cell(28, 3, "APPROVED", ln=True, align="C")
        pdf.set_font("Helvetica", "I", 5)
        pdf.set_xy(cx - 14, cy + 8)
        pdf.cell(28, 3, datetime.now().strftime("%Y-%m-%d"), ln=True, align="C")

    y_sig = pdf.get_y()

    # Left: Inspector
    pdf.set_draw_color(*GOLD)
    pdf.set_text_color(*DARK_TEXT)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_fill_color(*LIGHT_GRAY)
    pdf.rect(MARGIN, y_sig, 85, 24, "F")
    pdf.set_xy(MARGIN + 3, y_sig + 2)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_text_color(*NAVY)
    pdf.cell(0, 5, "Inspector", ln=True)
    pdf.set_xy(MARGIN + 3, y_sig + 8)
    pdf.set_text_color(*DARK_TEXT)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, "____________________________", ln=True)
    pdf.set_xy(MARGIN + 3, y_sig + 14)
    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(0, 4, "Kaustuv Guragain", ln=True)
    pdf.set_xy(MARGIN + 3, y_sig + 18)
    pdf.set_font("Helvetica", "I", 6)
    pdf.set_text_color(*GOLD)
    pdf.cell(0, 4, f"Date: {datetime.now().strftime('%Y-%m-%d')}", ln=True)

    # Center: Company stamp
    stamp_x = 115
    stamp_y = y_sig
    draw_stamp(stamp_x, stamp_y, 36)

    # Right: Reviewer
    rev_x = PAGE_W - MARGIN - 85
    rev_y = y_sig
    pdf.set_fill_color(*LIGHT_GRAY)
    pdf.rect(rev_x, rev_y, 85, 24, "F")
    pdf.set_xy(rev_x + 3, rev_y + 2)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_text_color(*NAVY)
    pdf.cell(0, 5, "Reviewed by", ln=True)
    pdf.set_xy(rev_x + 3, rev_y + 8)
    pdf.set_text_color(*DARK_TEXT)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, "____________________________", ln=True)
    pdf.set_xy(rev_x + 3, rev_y + 14)
    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(0, 4, "Nepal Trade Solutions", ln=True)
    pdf.set_xy(rev_x + 3, rev_y + 18)
    pdf.set_font("Helvetica", "I", 6)
    pdf.set_text_color(*GOLD)
    pdf.cell(0, 4, f"Date: {datetime.now().strftime('%Y-%m-%d')}", ln=True)

    pdf.set_y(max(y_sig + 24, stamp_y + 42) + 10)

    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(*GOLD)
    pdf.cell(0, 6, f"Report generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", ln=True)

    # ────────────────────────────────────────────────────
    # END OF REPORT
    # ────────────────────────────────────────────────────
    pdf.ln(10)
    pdf.set_draw_color(*GOLD)
    pdf.line(MARGIN, pdf.get_y(), PAGE_W - MARGIN, pdf.get_y())
    pdf.ln(3)
    pdf.set_text_color(*NAVY)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, "End Of The Report", ln=True, align="C")
    pdf.ln(2)
    pdf.set_text_color(*DARK_TEXT)
    pdf.set_font("Helvetica", "I", 7)
    pdf.multi_cell(0, 3.5, "Note: This report presents our observations and findings based on the inspection conducted at the specified location and time, using randomly selected samples. "
                 "The inspection was performed to the best of our professional judgment and ability. Our responsibility is limited to conducting a reasonable examination under standard inspection conditions. "
                 "This report does not exempt the supplier from their contractual obligations, nor does it affect the buyer's rights to claim compensation for any visible or hidden defects not identified during inspection or discovered thereafter. "
                 "This report is not a confirmation of shipment or dispatch. "
                 "All inspection services rendered by Nepal Trade Solutions Pvt. Ltd. are governed by our Terms and Conditions of Service, available upon request or via our official communication channels.")

    pdf.output(str(pdf_path))
    return str(pdf_path)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python report_generator.py <excel_path> [photo_dir]")
        sys.exit(1)
    excel = sys.argv[1]
    photos = sys.argv[2] if len(sys.argv) > 2 else "photos"
    result = build_report(excel, photos)
    print(f"Report generated: {result}")
